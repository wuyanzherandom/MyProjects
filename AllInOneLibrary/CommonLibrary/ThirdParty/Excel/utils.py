# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from django.db.models import fields as django_models_fields
from ..._import_lib import *


CELL_NAME = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
             'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
             'U', 'V', 'W', 'X', 'Y', 'Z',
             'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
             'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT',
             'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD',
             'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN',
             'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX',
             'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH',
             'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR',
             'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB',
             'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL',
             'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV',
             'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE', 'EF',
             'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP',
             'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ',
             'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ',
             'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR', 'FS', 'FT',
             'FU', 'FV', 'FW', 'FX', 'FY', 'FZ', 'GA', 'GB', 'GC', 'GD',
             'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN',
             'GO', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX',
             'GY', 'GZ', 'HA', 'HB', 'HC', 'HD', 'HE', 'HF', 'HG', 'HH',
             'HI', 'HJ', 'HK', 'HL', 'HM', 'HN', 'HO', 'HP', 'HQ', 'HR',
             'HS', 'HT', 'HU', 'HV', 'HW', 'HX', 'HY', 'HZ', 'IA', 'IB',
             'IC', 'ID', 'IE', 'IF', 'IG', 'IH', 'II', 'IJ', 'IK', 'IL',
             'IM', 'IN', 'IO', 'IP', 'IQ', 'IR', 'IS', 'IT', 'IU', 'IV',
             'IW', 'IX', 'IY', 'IZ', 'JA', 'JB', 'JC', 'JD', 'JE', 'JF',
             'JG', 'JH', 'JI', 'JJ', 'JK', 'JL', 'JM', 'JN', 'JO', 'JP',
             'JQ', 'JR', 'JS', 'JT', 'JU', 'JV', 'JW', 'JX', 'JY', 'JZ',
             'KA', 'KB', 'KC', 'KD', 'KE', 'KF', 'KG', 'KH', 'KI', 'KJ',
             'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ', 'KR', 'KS', 'KT',
             'KU', 'KV', 'KW', 'KX', 'KY', 'KZ', 'LA', 'LB', 'LC', 'LD',
             'LE', 'LF', 'LG', 'LH', 'LI', 'LJ', 'LK', 'LL', 'LM', 'LN',
             'LO', 'LP', 'LQ', 'LR', 'LS', 'LT', 'LU', 'LV', 'LW', 'LX',
             'LY', 'LZ', 'MA', 'MB', 'MC', 'MD', 'ME', 'MF', 'MG', 'MH',
             'MI', 'MJ', 'MK', 'ML', 'MM', 'MN', 'MO', 'MP', 'MQ', 'MR',
             'MS', 'MT', 'MU', 'MV', 'MW', 'MX', 'MY', 'MZ', 'NA', 'NB',
             'NC', 'ND', 'NE', 'NF', 'NG', 'NH', 'NI', 'NJ', 'NK', 'NL',
             'NM', 'NN', 'NO', 'NP', 'NQ', 'NR', 'NS', 'NT', 'NU', 'NV',
             'NW', 'NX', 'NY', 'NZ', 'OA', 'OB', 'OC', 'OD', 'OE', 'OF',
             'OG', 'OH', 'OI', 'OJ', 'OK', 'OL', 'OM', 'ON', 'OO', 'OP',
             'OQ', 'OR', 'OS', 'OT', 'OU', 'OV', 'OW', 'OX', 'OY', 'OZ',
             'PA', 'PB', 'PC', 'PD', 'PE', 'PF', 'PG', 'PH', 'PI', 'PJ',
             'PK', 'PL', 'PM', 'PN', 'PO', 'PP', 'PQ', 'PR', 'PS', 'PT',
             'PU', 'PV', 'PW', 'PX', 'PY', 'PZ', 'QA', 'QB', 'QC', 'QD',
             'QE', 'QF', 'QG', 'QH', 'QI', 'QJ', 'QK', 'QL', 'QM', 'QN',
             'QO', 'QP', 'QQ', 'QR', 'QS', 'QT', 'QU', 'QV', 'QW', 'QX',
             'QY', 'QZ', 'RA', 'RB', 'RC', 'RD', 'RE', 'RF', 'RG', 'RH',
             'RI', 'RJ', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RQ', 'RR',
             'RS', 'RT', 'RU', 'RV', 'RW', 'RX', 'RY', 'RZ', 'SA', 'SB',
             'SC', 'SD', 'SE', 'SF', 'SG', 'SH', 'SI', 'SJ', 'SK', 'SL',
             'SM', 'SN', 'SO', 'SP', 'SQ', 'SR', 'SS', 'ST', 'SU', 'SV',
             'SW', 'SX', 'SY', 'SZ', 'TA', 'TB', 'TC', 'TD', 'TE', 'TF',
             'TG', 'TH', 'TI', 'TJ', 'TK', 'TL', 'TM', 'TN', 'TO', 'TP',
             'TQ', 'TR', 'TS', 'TT', 'TU', 'TV', 'TW', 'TX', 'TY', 'TZ',
             'UA', 'UB', 'UC', 'UD', 'UE', 'UF', 'UG', 'UH', 'UI', 'UJ',
             'UK', 'UL', 'UM', 'UN', 'UO', 'UP', 'UQ', 'UR', 'US', 'UT',
             'UU', 'UV', 'UW', 'UX', 'UY', 'UZ', 'VA', 'VB', 'VC', 'VD',
             'VE', 'VF', 'VG', 'VH', 'VI', 'VJ', 'VK', 'VL', 'VM', 'VN',
             'VO', 'VP', 'VQ', 'VR', 'VS', 'VT', 'VU', 'VV', 'VW', 'VX',
             'VY', 'VZ', 'WA', 'WB', 'WC', 'WD', 'WE', 'WF', 'WG', 'WH',
             'WI', 'WJ', 'WK', 'WL', 'WM', 'WN', 'WO', 'WP', 'WQ', 'WR',
             'WS', 'WT', 'WU', 'WV', 'WW', 'WX', 'WY', 'WZ', 'XA', 'XB',
             'XC', 'XD', 'XE', 'XF', 'XG', 'XH', 'XI', 'XJ', 'XK', 'XL',
             'XM', 'XN', 'XO', 'XP', 'XQ', 'XR', 'XS', 'XT', 'XU', 'XV',
             'XW', 'XX', 'XY', 'XZ', 'YA', 'YB', 'YC', 'YD', 'YE', 'YF',
             'YG', 'YH', 'YI', 'YJ', 'YK', 'YL', 'YM', 'YN', 'YO', 'YP',
             'YQ', 'YR', 'YS', 'YT', 'YU', 'YV', 'YW', 'YX', 'YY', 'YZ',
             'ZA', 'ZB', 'ZC', 'ZD', 'ZE', 'ZF', 'ZG', 'ZH', 'ZI', 'ZJ',
             'ZK', 'ZL', 'ZM', 'ZN', 'ZO', 'ZP', 'ZQ', 'ZR', 'ZS', 'ZT',
             'ZU', 'ZV', 'ZW', 'ZX', 'ZY', 'ZZ']


class Export(View):
    model = None
    fields = []
    display_fields = []
    type = {}
    filters = {}
    filename = None

    def get(self, request, *args, **kwargs):
        if self.filters:
            return HttpResponse(create_excel(
                filename1=self.filename if self.filename else 'Export ' + self.model.__name__ + ' (' + datetime.datetime.now().strftime("%Y%m%d") + ')',
                query=self.model.objects.filter(*self.filters).values(*self.fields),
                fields=self.fields,
                save_internal=False,
                display_fields=self.display_fields,
                types=self.type,
            ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            return HttpResponse(create_excel(
                filename1=self.filename if self.filename else 'Export ' + self.model.__name__ + ' (' + datetime.datetime.now().strftime("%Y%m%d") + ')',
                query=self.model.objects.all().values(*self.fields),
                fields=self.fields,
                save_internal=False,
                display_fields=self.display_fields,
                types=self.type,
            ), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def create_excel(query, fields, display_fields, types=None, filename1=r'test1.xlsx', save_internal=True):
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    wb = Workbook()
    ws = wb.active
    for count, key in enumerate(display_fields):
        ws[CELL_NAME[count]+'1'] = key
    for count1, record in enumerate(query):
        for count2, key in enumerate(fields):
            if types:
                _type = types.get(key)
                if _type:
                    if _type['type'] == 'map':
                        try:
                            record[key] = _type['map'][record[key]]
                        except:
                            record[key] = _type.get('default', None)
                    elif _type['type'] == 'lookup':
                        try:
                            value = _type['models'].objects.filter(
                                id=record[key]
                            ).values(_type['field'])
                            record[key] = value[0][_type['field']]
                        except:
                            record[key] = None
            ws[CELL_NAME[count2]+str(count1+2)] = record[key]
    if save_internal:
        wb.save(filename=filename1)
    else:  # save_internal - not return to HTTP response
        return save_virtual_workbook(wb)
        #except:
            #return "invalid query format or save error"


"""
Example:
ImportF(
        request=request,
        fields=[
            "name", "date", "employee", "in_time",
            "break_time1", "resume_time1", "break_time2", "resume_time2",
            "out_time",
        ],
        display_fields=[
            "Name", "Date", "Employee", "In time",
            "Break Time1", "Resume Time1", "Break Time2", "Resume Time2",
            "Out time",
        ],
        types={
            "employee": {
                "type": "lookup",
                "models": models.Employee,
                "field": "name",
            },
        },
        model=models.TimeTableRecord,
        check_fields=[],
    )
"""
def ImportF(display_fields, fields, types, model, check_fields, exclude_fields=None, preSave=None, postSave=None, request=None, filename=None):
    from ...operation.list.utils import _check_key_list, _compare_list_with_display_fields

    if (not request and not filename) or (request and filename):
        return {'flag': False, 'info': '', 'msg': 'either request or filename can be selected'}

    if request:
        try:
            wb = load_workbook(request.FILES['xls'])
        except:
            return {'flag': False, 'info': '', 'msg': 'excel file is not found'}
    if filename:
        try:
            wb = load_workbook(filename)
        except:
            return {'flag': False, 'info': '', 'msg': 'xls'}
    ws = wb.active
    count_row = 0
    _create, _new, _update, _exists = 0, 0, 0, 0
    _new_index = {}
    needed_field = []
    default_values = []
    null_when_blanks = []
    field = model._meta.fields
    for f in field:
        if not getattr(f, 'blank', False):
            if getattr(f, 'null', False):
                null_when_blanks.append(f.name)
            else:
                needed_field.append(f.name)

        default_value = getattr(f, 'default')
        if default_value and default_value != django_models_fields.NOT_PROVIDED:
            default_values.append({"name": f.name, "value": default_value})
    for count, row in enumerate(ws.iter_rows()):
        if len(row) > 0:
            if count == 0:
                get_lists = [row[x].value for x in xrange(ws.max_column)]  # get 1st row of value
                flag, info, msg = _check_key_list(
                    _lists=get_lists,
                    _lists_keys_needed=display_fields,
                )
                if not flag:
                    return {'flag': flag, 'info': info, 'msg': msg}
                flag, info, msg = _check_key_list(
                    _lists=fields,
                    _lists_keys_needed=needed_field,
                )
                if not flag:
                    return {'flag': flag, 'info': info, 'msg': msg}
                _new_index = _compare_list_with_display_fields(
                    lists=get_lists,
                    display_fields=display_fields,
                    fields=fields,
                )
                if not _new_index:
                    return {'flag': False, 'info': '', 'msg': 'some field is needed'}
            else:
                record = dict()
                if 'user_id' in model._meta.get_fields():
                    record['user_id'] = request.user.id
                for field in _new_index:
                    _type = types.get(field)
                    if _type:
                        if _type.get('trim') == True:
                            record[field] = unicode(row[_new_index[field]].value).strip()

                        if _type.get('type') == 'map':
                            try:
                                record[field] = _type['map'][unicode(row[_new_index[field]].value).strip()]
                            except:
                                record[field] = _type['default']
                        elif _type.get('type') == 'lookup':
                            value = {_type['field']: str(unicode(row[_new_index[field]].value).strip())}
                            try:
                                r = _type['models'].objects.get(
                                    **value
                                )
                                record[field+'_id'] = r.pk
                            except:
                                record[field] = None
                        elif _type.get('type') == 'date':
                            try:
                                record[field] = row[_new_index[field]].value.date()
                            except:
                                record[field] = None
                    else:
                        record[field] = row[_new_index[field]].value

                for default_value in default_values:
                    if not record.get(default_value['name']):
                        record[default_value['name']] = default_value['value']

                for null_when_blank in null_when_blanks:
                    if not null_when_blank.name in record or record.get(null_when_blank.name) == '':
                        record[null_when_blank.name] = None
                """
                obj = model.objects.only('pk').get(**check_fields_dict)
                model.objects.filter(pk=obj.pk).update(**record)
                _exists += 1
                """
                if exclude_fields:
                    filtered_record = {k: v for k, v in record.iteritems() if k not in exclude_fields}
                else:
                    filtered_record = record

                check_fields_dict = {}
                for x in check_fields:
                    check_fields_dict[x] = record[x]

                get_obj = model.objects.filter(
                    **filtered_record
                )

                if check_fields_dict:
                    get_obj = get_obj.filter(**check_fields_dict)
                else:
                    get_obj = get_obj

                if len(get_obj) == 0:  # create new
                    if preSave:
                        preSave(record, None)
                    #raise Exception(filtered_record)
                    obj = model.objects.create(
                        **filtered_record
                    )
                    if postSave:
                        postSave(record, obj)
                    _new += 1

                elif len(get_obj) == 1:  # update
                    if check_fields:
                        for check_field in check_fields:
                            record.pop(check_field)
                        if preSave:
                            preSave(record, get_obj[0])
                        for attr in filtered_record:
                            if attr not in check_fields:
                                setattr(get_obj[0], attr, record[attr])
                        get_obj[0].save()
                        if postSave:
                            postSave(record, get_obj[0])
                        _update += 1
                    _exists += 1
    return {'total': count_row, 'new': _new, 'update': _update, 'exists': _exists}
